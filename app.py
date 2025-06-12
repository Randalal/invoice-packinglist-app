import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from datetime import datetime
from io import BytesIO

# 设置页面标题与布局
st.set_page_config(page_title="Invoice & Packing List Generator", layout="centered")
st.title("Invoice & Packing List Generator")

# -------- Step 1: Upload  --------

# Step 1.1: 上传模板文件（用于预览和后续字段写入）
st.header("Step 1: Upload Invoice or Packing List Template")

# 上传控件（合并预览与写入用途）
uploaded_template = st.file_uploader("Click to pload invoice or packing list template (.xlsx)", type=["xlsx"], key="template_upload")

if uploaded_template:
    try:
        # 预览上传内容
        df = pd.read_excel(uploaded_template)
        st.success("Template uploaded successfully.")
        st.markdown("Preview of the uploaded file:")
        st.dataframe(df.head())

        # 保存为字节流，供 openpyxl 使用（后续步骤写入字段用）
        st.session_state["invoice_template_file"] = uploaded_template.getvalue()

    except Exception as e:
        st.error(f"Failed to process the uploaded file: {str(e)}")

# -------- Step 1.2: Upload PI File --------
# Step 1.2: 上传 PI 文件（Proforma Invoice）
st.header("Step 2: Upload PI File")

uploaded_pi = st.file_uploader("Click to upload PI file", type=["xlsx"], key="pi_upload")

if uploaded_pi:
    try:
        # 加载工作簿与工作表
        pi_wb = load_workbook(filename=BytesIO(uploaded_pi.read()), data_only=True)
        pi_ws = pi_wb.active

        # 提取 D10 ~ D13 的 "Billing To" 信息
        bill_to_lines = []
        for row in range(10, 14):
            value = pi_ws[f"D{row}"].value
            if value:
                bill_to_lines.append(str(value))

        if bill_to_lines:
            st.markdown("Billing To Information:")
            for line in bill_to_lines:
                st.markdown(line)
        else:
            st.warning("No 'Billing To' information found in cells D10 to D13.")

        # 提取第14行起的产品信息（最多提取20行）
        def get_merged_cell_value(ws, row, col):
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    return ws.cell(row=min_row, column=min_col).value
            return ws.cell(row=row, column=col).value

        pi_data = []
        row_index = 14
        for _ in range(20):
            ean = get_merged_cell_value(pi_ws, row_index, 3)
            desc = get_merged_cell_value(pi_ws, row_index, 6)
            qty = get_merged_cell_value(pi_ws, row_index, 9)
            price = get_merged_cell_value(pi_ws, row_index, 10)
            total = get_merged_cell_value(pi_ws, row_index, 13)

            if all(x is None for x in [ean, desc, qty, price, total]):
                break

            pi_data.append({
                "EAN": ean,
                "Description": desc,
                "Quantity": qty,
                "Price (USD)": price,
                "Total (USD)": total
            })
            row_index += 1

        if pi_data:
            st.success(f"PI file uploaded and parsed successfully: {uploaded_pi.name}")
            st.dataframe(pd.DataFrame(pi_data))
        else:
            st.warning("No valid product data found from row 14 onward.")

        # 存入 session_state 供 Step 5 使用
        st.session_state["bill_to_lines"] = bill_to_lines
        st.session_state["product_list"] = pi_data

    except Exception as e:
        st.error(f"Failed to process PI file: {str(e)}")

# -------- Step 1.3: Upload Product List --------
# Step 1.3: Upload Product List File
st.header("Step 3: Upload Product List File")

uploaded_product = st.file_uploader("Click to upload product list file", type=["xlsx"], key="product_upload")

if uploaded_product:
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # 加载工作簿与工作表
        product_wb = load_workbook(filename=BytesIO(uploaded_product.read()), data_only=True)
        product_ws = product_wb.active

        # 读取第 3~7 行内容
        product_data = []
        for i in range(3, 8):
            ean = product_ws[f"A{i}"].value
            name = product_ws[f"B{i}"].value
            qty = product_ws[f"C{i}"].value

            if all(x is None for x in [ean, name, qty]):
                break  # 跳过空行

            product_data.append({
                "EAN": ean,
                "Description": name,   # 为兼容 Step 5 格式
                "Quantity": qty,
                "Price (USD)": None,
                "Total (USD)": None
            })

        if product_data:
            st.success(f"Product list uploaded successfully: {uploaded_product.name}")
            st.markdown("Preview of the product list:")
            st.dataframe(pd.DataFrame(product_data))
        else:
            st.warning("No valid product data found in rows 3 to 7.")

        # 存入 session_state
        st.session_state["product_list"] = product_data

    except Exception as e:
        st.error(f"Failed to process product list file: {str(e)}")

# -------- Step 1.4: Upload Packing Sheet --------
# Step 1.4: Upload Packing Sheet
st.header("Step 4: Upload Packing Sheet")

uploaded_packing = st.file_uploader("Click to upload packing sheet", type=["xlsx"], key="packing_upload")

if uploaded_packing:
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        import pandas as pd

        wb = load_workbook(filename=BytesIO(uploaded_packing.read()), data_only=True)
        ws = wb.active

        invoice_no = ws["B2"].value
        billing_party = ws["E3"].value
        consignee = ws["B4"].value
        date = ws["H2"].value

        st.success("Packing sheet uploaded successfully.")

        st.markdown(f"**Invoice Number:** {invoice_no}")
        st.markdown(f"**Billing Party:** {billing_party}")
        st.markdown(f"**Consignee:** {consignee}")
        st.markdown(f"**Date:** {date}")

        # 读取产品明细（从第 7 行开始）
        product_rows = []
        for row in ws.iter_rows(min_row=7, max_col=8, values_only=True):
            sn, item_code, desc, qty, pkg, weight, dimension = row[:7]
            if all(x is None for x in [sn, item_code, desc, qty]):
                break
            product_rows.append({
                "SN Number": sn,
                "Item Code": item_code,
                "Description": desc,
                "Quantity": qty,
                "Package": pkg,
                "Weight": weight,
                "Dimension": dimension
            })

        if product_rows:
            st.markdown("Preview of packed items:")
            st.dataframe(pd.DataFrame(product_rows))
        else:
            st.warning("No valid packed items found.")

        # 包装信息（E12、F12、G12）
        pkg_type = ws["E12"].value
        weight = ws["F12"].value
        dimension = ws["G12"].value

        st.markdown("Package Info:")
        st.markdown(f"- Package Type: {pkg_type}")
        st.markdown(f"- Weight: {weight}")
        st.markdown(f"- Dimension: {dimension}")

        # 存储 package_code（用于 Step 5）
        st.session_state["package_code"] = pkg_type

        # Sheet2 信息（收货人地址、电话等）
        if "Sheet2" in wb.sheetnames:
            sheet2 = wb["Sheet2"]
            try:
                ref = sheet2["A5"].value
                name = sheet2["B5"].value
                company = sheet2["C5"].value
                address = sheet2["D5"].value
                city = sheet2["E5"].value
                phone = sheet2["F5"].value
                country = sheet2["G5"].value
                awb = sheet2["H5"].value
                code = sheet2["I5"].value

                st.markdown("Shipping Info (from Sheet2):")
                st.markdown(f"- Reference: {ref}")
                st.markdown(f"- Recipient: {name} ({company})")
                st.markdown(f"- Address: {address}, {city}, {country}")
                st.markdown(f"- Phone: {phone}")
                st.markdown(f"- AWB: {awb}")

                st.session_state["sheet2_info"] = {
                    "D5": address,
                    "E5": city,
                    "F5": phone,
                    "G5": country,
                    "H5": awb,
                    "I5": code
                }

            except Exception as e:
                st.warning(f"Failed to extract shipping info from Sheet2: {str(e)}")
        else:
            st.warning("Sheet2 not found in uploaded file.")

    except Exception as e:
        st.error(f"Failed to read packing sheet: {str(e)}")

# -------- Step 1.5: Upload HS Code File --------
st.header("Step 5: Upload HS Code File")

uploaded_hs = st.file_uploader("Click to upload HS Code file", type=["xlsx"], key="hs_upload")

if uploaded_hs:
    try:
        hs_wb = load_workbook(filename=BytesIO(uploaded_hs.read()), data_only=True)
        hs_ws = hs_wb.active

        hs_mapping = {}
        for row in hs_ws.iter_rows(min_row=2, max_col=2, values_only=True):
            ean, hs_code = row
            if ean and hs_code:
                hs_mapping[str(ean).strip()] = str(hs_code).strip()

        if hs_mapping:
            st.success(f"HS Code file uploaded and parsed successfully: {uploaded_hs.name}")
            st.write(f"{len(hs_mapping)} rows loaded.")
            st.dataframe(pd.DataFrame(list(hs_mapping.items()), columns=["EAN", "HS Code"]))
            st.session_state["hs_mapping"] = hs_mapping
        else:
            st.warning("No valid HS Code mappings found in the uploaded file.")

    except Exception as e:
        st.error(f"Failed to process HS Code file: {str(e)}")


# -------- Step 2: Fill Invoice Template --------

def fill_invoice_template_core():
    try:
        invoice_template_file = st.session_state["invoice_template_file"]
        wb = load_workbook(filename=BytesIO(invoice_template_file))
        ws = wb["Invoice"] if "Invoice" in wb.sheetnames else wb.active

        # Step 2.1 - 填写 BILL TO
        bill_to_lines = st.session_state.get("bill_to_lines", [])
        for i, line in enumerate(bill_to_lines):
            ws[f"A{14 + i}"] = line

        # Step 2.2 - 填写收件人地址信息
        info = st.session_state.get("sheet2_info", {})
        ws["A19"] = info.get("E5", "")
        ws["A20"] = info.get("F5", "")
        ws["A21"] = f"{info.get('D5', '')} {info.get('H5', '')}"
        ws["A22"] = f"{info.get('G5', '')}, {info.get('I5', '')}"
        ws["F16"] = f"{info.get('G5', '')}, {info.get('I5', '')}"

        # Step 2.3 - 包装类型、日期、原产地
        # 获取包装类型并格式化
        raw_code = st.session_state.get("package_code", "")  # e.g., "01 # PLT"
        parts = raw_code.strip().split()

        if parts and parts[0].isdigit():
            count = int(parts[0])
            ws["A24"] = f"{count} Pallet(s)"
        else:
            ws["A24"] = raw_code  # fallback 原始值

        # 当前日期填入 F10，格式为 DD/MM/YYYY
        ws["F10"] = datetime.today().strftime("%d/%m/%Y")

        # 原产地写入 G24
        ws["G24"] = "Made in China"

        # Step 2.4: 插入产品明细（包括价格查找与文本格式）

        product_list = st.session_state.get("product_list", [])
        pi_file = st.session_state.get("pi_upload", None)

        # 读取 PI 文件（用于价格匹配）
        pi_data = {}
        pi_number = ""

        if pi_file:
            wb_pi = load_workbook(filename=BytesIO(pi_file.getvalue()), data_only=True)
            ws_pi = wb_pi.active

            # 获取 PI number from K6
            pi_number = ws_pi["K6"].value or ""

            # 遍历 PI 表格中内容，构建 EAN => {单价，总价}
            for row in ws_pi.iter_rows(min_row=12, max_row=100, values_only=True):
                ean = str(row[1]).strip() if row[1] else ""
                unit_price = row[9] if len(row) > 9 else None
                total_price = row[11] if len(row) > 11 else None
                if ean and unit_price is not None and total_price is not None:
                    pi_data[ean] = {
                        "unit_price": unit_price,
                        "total_price": total_price
                    }

        # 插入产品明细，从第26行开始，向下插入 len(product_list) 行
        start_row = 26
        ws.insert_rows(start_row, amount=len(product_list))

        for i, prod in enumerate(product_list):
            row = start_row + i
            ean = str(prod.get("EAN", "")).strip()
            desc = prod.get("Description", "").strip()

            # B列：EAN，设置为文本格式防止科学计数
            cell = ws[f"B{row}"]
            cell.value = ean
            cell.number_format = "@"

            ws[f"C{row}"] = pi_number  # PI number
            ws[f"D{row}"] = desc  # Description
            ws[f"E{row}"] = "China"  # Origin
            ws[f"F{row}"] = prod.get("Quantity", "")  # PCS

            # G/H：从 PI 文件中获取价格
            if ean in pi_data:
                ws[f"G{row}"] = pi_data[ean]["unit_price"]
                ws[f"H{row}"] = pi_data[ean]["total_price"]
            else:
                ws[f"G{row}"] = ""
                ws[f"H{row}"] = ""

        # 居中 + 边框
        from openpyxl.styles import Alignment, Border, Side

        # 样式设置：居中 + 黑色细边框
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        for i, prod in enumerate(product_list):
            row = start_row + i
            ean = str(prod.get("EAN", "")).strip()
            desc = prod.get("Description", "").strip()

            # B列：EAN，文本格式 + 样式
            cell = ws[f"B{row}"]
            cell.value = ean
            cell.number_format = "@"
            cell.alignment = center_align
            cell.border = thin_border

            # C 到 H 列：赋值 + 样式
            col_map = {
                "C": pi_number,
                "D": desc,
                "E": "China",
                "F": prod.get("Quantity", "")
            }

            # G/H 列来自 PI 文件
            if ean in pi_data:
                col_map["G"] = pi_data[ean]["unit_price"]
                col_map["H"] = pi_data[ean]["total_price"]
            else:
                col_map["G"] = ""
                col_map["H"] = ""

            # 写入 C~H 列，统一加样式
            for col, val in col_map.items():
                cell = ws[f"{col}{row}"]
                cell.value = val
                cell.alignment = center_align
                cell.border = thin_border

        # Step 2.5: 添加 Total 汇总行 -------
        from openpyxl.styles import Alignment, Border, Side

        # 居中样式
        center_align = Alignment(horizontal="center", vertical="center")

        # 设置边框样式（仅用于产品数据行）
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 1. 设置产品行格式：居中 + 边框 + 不折叠
        for i in range(len(product_list)):
            row = start_row + i
            for col in ["B", "C", "D", "E", "F", "G", "H"]:
                cell = ws[f"{col}{row}"]
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[row].outlineLevel = 0
            ws.row_dimensions[row].hidden = False

        # 2. 动态查找 Total 行（只写入数值，不改边框）
        total_row = None
        for row in range(start_row + len(product_list), start_row + len(product_list) + 10):
            if ws[f"D{row}"].value and str(ws[f"D{row}"].value).strip().lower() == "total":
                total_row = row
                break

        if total_row:
            # 计算总值
            total_pcs = sum(int(prod.get("Quantity", 0)) for prod in product_list)
            total_usd = sum(
                float(pi_data.get(str(prod.get("EAN", "")).strip(), {}).get("total_price", 0))
                for prod in product_list
            )

            ws[f"F{total_row}"] = total_pcs
            ws[f"H{total_row}"] = total_usd

            ws[f"F{total_row}"].alignment = center_align
            ws[f"H{total_row}"].alignment = center_align

            # Total 行也不折叠
            ws.row_dimensions[total_row].outlineLevel = 0
            ws.row_dimensions[total_row].hidden = False
        else:
            st.warning("模板中找不到 Total 行，请确认 D 列是否包含 'Total'")

        # ------- Step 2.6: 根据 EAN 写入 HS Code 到 A 列 -------
        hs_mapping = st.session_state.get("hs_mapping", {})
        product_list = st.session_state.get("product_list", [])

        start_row = 26

        from openpyxl.styles import Alignment, Border, Side

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for i, prod in enumerate(product_list):
            row = start_row + i
            ean = str(prod.get("EAN", "")).strip()
            hs_code = hs_mapping.get(ean, "")

            cell = ws[f"A{row}"]
            cell.value = str(hs_code)  # 明确转为字符串
            cell.number_format = "@"  # 设置为文本格式
            cell.alignment = center_align
            cell.border = thin_border

        # 保存到内存
        output_buffer = BytesIO()
        wb.save(output_buffer)
        st.session_state["final_invoice_file"] = output_buffer

        return True, "Invoice and Packing list filled successfully."

    except Exception as e:
        return False, f"Error during invoice core filling: {str(e)}"



# -------- Fill Invoice Template --------

st.subheader("Fill Invoice and Packing list Info")

if st.button("Fill Invoice Template"):
    success, msg = fill_invoice_template_core()
    if success:
        st.success(msg)
    else:
        st.error(msg)

    st.markdown("Session keys:")
    st.write(list(st.session_state.keys()))


# -------- Export Final Invoice --------

if "final_invoice_file" in st.session_state:
    st.download_button(
        label="Download Final Invoice File",
        data=st.session_state["final_invoice_file"].getvalue(),
        file_name="Generated_Invoice.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Invoice file not ready. Please complete Step 5 first.")

